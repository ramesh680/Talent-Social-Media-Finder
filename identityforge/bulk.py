"""
Bulk ingest for CSV / XLSX.

Design constraints that shaped this, all of them real:

1. HEADERS ARE NEVER WHAT YOU SPECIFIED. A sheet that came from a client, a
   Zendesk export or a colleague will say "Talent Name" or "title" or
   "Full Name", not "name". Rejecting the file over a header is the fastest way
   to make a tool unused, so headers are matched tolerantly and the mapping is
   reported back so the operator can see how each column was read.

2. A BULK RUN CANNOT FINISH INSIDE AN HTTP REQUEST. Live resolution costs
   roughly 4-10 upstream calls per name, and the rate limiter deliberately
   paces MusicBrainz at 1.1s and Wikidata at 0.4s. That is ~5-10 seconds per
   name, against a 120s gunicorn timeout. So this processes under a wall-clock
   budget and returns the unprocessed rows rather than dying at row 14. The
   operator resubmits the remainder; the HTTP cache makes the second pass
   cheaper.

3. PARTIAL SUCCESS MUST BE LEGIBLE. Every output row carries its decision and
   its evidence, so 'resolved', 'needs a human', and 'not attempted' are three
   visibly different states, not three flavours of blank.
"""

from __future__ import annotations

import csv
import io
import time
from dataclasses import dataclass, field
from typing import Callable, Iterable, Optional

from .authorities import OCCUPATION_BUCKETS
from .platforms import TARGET_PLATFORMS, Platform

# ---------------------------------------------------------------------------
# schema
# ---------------------------------------------------------------------------

INPUT_COLUMNS = ["row_id", "name", "role", "active_year", "country",
                 "context", "client", "notes"]
REQUIRED_COLUMNS = ["name"]

PLATFORM_ORDER = [Platform.FACEBOOK, Platform.INSTAGRAM, Platform.TWITTER,
                  Platform.YOUTUBE, Platform.WIKIPEDIA, Platform.IMDB,
                  Platform.TIKTOK, Platform.LINKEDIN]

OUTPUT_COLUMNS = (
    ["row_id", "name", "decision", "entity_id", "label", "roles",
     "coverage", "label_match"]
    + [p.value for p in PLATFORM_ORDER]
    + ["needs_review", "spoke_ids", "alternates", "notes"]
)

# Header aliases. Lowercased, punctuation stripped, spaces collapsed.
_ALIASES: dict[str, str] = {}


def _norm_header(raw) -> str:
    """The single normalisation used for BOTH registration and lookup.

    Registering 'row_id' while lookup normalises to 'row id' silently loses the
    alias, so both sides must go through this one function.
    """
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    for ch in ("_", "-", ".", "/", "\\"):
        s = s.replace(ch, " ")
    return " ".join(s.split())


def _register(canonical: str, *aliases: str) -> None:
    for a in (canonical,) + aliases:
        k = _norm_header(a)
        if k:
            _ALIASES[k] = canonical


_register("name", "talent name", "talent", "full name", "fullname", "title",
          "person", "person name", "celebrity", "influencer", "creator",
          "artist", "artist name", "profile name", "entity", "entity name")
_register("role", "expected role", "category", "title category",
          "title_category", "occupation", "profession", "type", "vertical",
          "sub category", "title sub category")
_register("active_year", "year", "active year", "release year", "campaign year",
          "reference year")
_register("country", "citizenship", "nationality", "market", "region",
          "country qid")
_register("context", "show", "brand", "campaign", "programme", "program",
          "series", "project", "note context", "source")
_register("client", "account", "customer", "brand set", "brandset")
_register("row_id", "id", "row", "sr no", "srno", "s no", "sno", "index",
          "ticket", "ticket id", "request id")
_register("notes", "note", "comment", "comments", "remarks")


def canon_header(raw: str, strict: bool = False) -> Optional[str]:
    """
    Map a spreadsheet header to a canonical field, or None if unrecognised.

    strict=True disables the substring fallback. Use it when DETECTING which
    row is the header: a title cell like "Talent list for Q3" contains the
    token "talent" and would otherwise be read as a name column, making the
    title row the header and the subtitle row the first record.
    """
    k = _norm_header(raw)
    if not k:
        return None
    if k in _ALIASES:
        return _ALIASES[k]
    if strict:
        return None
    # a real header may be decorated: "Talent Name (English)"
    if len(k) > 60:
        return None                 # a sentence, not a column heading
    # longest alias first, so 'talent name' beats 'talent'
    for alias, canonical in sorted(_ALIASES.items(), key=lambda kv: -len(kv[0])):
        if len(alias) > 3 and alias in k:
            return canonical
    return None


def score_header_row(row: list) -> int:
    """How many cells in `row` are recognisable column headings, strictly."""
    seen = set()
    for cell in row:
        c = canon_header(cell, strict=True)
        if c:
            seen.add(c)
    return len(seen)


# Country names -> Wikidata Q-ids, for the markets that actually appear in
# ListenFirst work. An unrecognised value is passed through untouched so a
# raw Q-id still works.
COUNTRY_QIDS = {
    "india": "Q668", "united states": "Q30", "usa": "Q30", "us": "Q30",
    "united kingdom": "Q145", "uk": "Q145", "canada": "Q16",
    "australia": "Q408", "japan": "Q17", "south korea": "Q884",
    "korea": "Q884", "france": "Q142", "germany": "Q183", "spain": "Q29",
    "italy": "Q38", "brazil": "Q155", "mexico": "Q96", "nigeria": "Q1033",
    "indonesia": "Q252", "philippines": "Q928", "thailand": "Q869",
    "china": "Q148", "russia": "Q159", "netherlands": "Q55", "sweden": "Q34",
    "uae": "Q878", "saudi arabia": "Q851", "singapore": "Q334",
}


def canon_country(value: str) -> str:
    v = (value or "").strip()
    if not v:
        return ""
    if v.upper().startswith("Q") and v[1:].isdigit():
        return v.upper()
    return COUNTRY_QIDS.get(v.lower(), v)


def canon_role(value: str) -> tuple[str, str]:
    """Return (role, warning). Unknown roles are dropped, not guessed."""
    v = " ".join((value or "").strip().lower().split())
    if not v:
        return "", ""
    if v in OCCUPATION_BUCKETS:
        return v, ""
    synonyms = {
        "singer": "musician", "music": "musician", "musician/singer": "musician",
        "composer": "musician", "rapper": "musician", "band": "musician",
        "actor": "actor", "actress": "actor", "film": "actor", "tv": "actor",
        "movie": "actor", "director": "director", "filmmaker": "director",
        "producer": "director", "athlete": "athlete", "sports": "athlete",
        "sport": "athlete", "cricketer": "athlete", "footballer": "athlete",
        "youtuber": "creator", "streamer": "creator", "influencer": "creator",
        "content creator": "creator", "digital creator": "creator",
        "model": "model", "journalist": "journalist", "anchor": "journalist",
        "politician": "politician", "writer": "writer", "author": "writer",
        "ceo": "executive", "founder": "executive", "executive": "executive",
        "comedian": "comedian", "chef": "chef",
    }
    if v in synonyms:
        return synonyms[v], ""
    return "", f"unrecognised role {value!r} - treated as unknown"


# ---------------------------------------------------------------------------
# parsing
# ---------------------------------------------------------------------------

@dataclass
class ParsedRow:
    row_number: int
    name: str = ""
    role: str = ""
    active_year: Optional[int] = None
    country: str = ""
    context: str = ""
    client: str = ""
    notes: str = ""
    warnings: list[str] = field(default_factory=list)
    row_id: str = ""

    @property
    def valid(self) -> bool:
        return bool(self.name.strip())


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    header_map: dict[str, str] = field(default_factory=dict)
    unmapped_headers: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def valid_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if r.valid]

    @property
    def invalid_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if not r.valid]


def _coerce_year(v) -> tuple[Optional[int], str]:
    s = str(v or "").strip()
    if not s:
        return None, ""
    s = s.split(".")[0]                     # Excel gives 2005.0
    if s.isdigit() and 1800 <= int(s) <= 2100:
        return int(s), ""
    return None, f"ignored active_year {v!r}"


def _build_rows(header: list, data_rows: Iterable[list]) -> ParseResult:
    res = ParseResult()
    mapping: dict[int, str] = {}
    for i, h in enumerate(header):
        c = canon_header(h)
        if c:
            # first column wins, so a stray later match cannot hijack a field
            if c not in mapping.values():
                mapping[i] = c
                res.header_map[str(h)] = c
        elif str(h or "").strip():
            res.unmapped_headers.append(str(h).strip())

    if "name" not in mapping.values():
        res.errors.append(
            "No name column found. Add a column headed 'name' (or Talent Name, "
            "Full Name, Title). Headers seen: "
            + ", ".join(str(h) for h in header if str(h or "").strip()))
        return res

    for n, raw in enumerate(data_rows, start=2):
        if not any(str(c or "").strip() for c in raw):
            continue                        # blank row
        pr = ParsedRow(row_number=n)
        for idx, field_name in mapping.items():
            val = raw[idx] if idx < len(raw) else ""
            val = "" if val is None else str(val).strip()
            if field_name == "name":
                pr.name = val
            elif field_name == "role":
                pr.role, w = canon_role(val)
                if w:
                    pr.warnings.append(w)
            elif field_name == "active_year":
                pr.active_year, w = _coerce_year(val)
                if w:
                    pr.warnings.append(w)
            elif field_name == "country":
                pr.country = canon_country(val)
            elif field_name == "context":
                pr.context = val
            elif field_name == "client":
                pr.client = val
            elif field_name == "notes":
                pr.notes = val
            elif field_name == "row_id":
                pr.row_id = val
        if not pr.row_id:
            pr.row_id = str(n - 1)
        if not pr.valid:
            pr.warnings.append("empty name - skipped")
        res.rows.append(pr)
    return res


def parse_csv(data: bytes) -> ParseResult:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        r = ParseResult()
        r.errors.append("Could not decode the file as UTF-8 or Latin-1.")
        return r
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = list(csv.reader(io.StringIO(text), dialect))
    if not reader:
        r = ParseResult()
        r.errors.append("The file is empty.")
        return r
    return _build_rows(reader[0], reader[1:])


def parse_xlsx(data: bytes) -> ParseResult:
    try:
        from openpyxl import load_workbook
    except ImportError:                                   # pragma: no cover
        r = ParseResult()
        r.errors.append("openpyxl is not installed - cannot read .xlsx.")
        return r
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:                              # noqa: BLE001
        r = ParseResult()
        r.errors.append(f"Could not open the workbook: {exc}")
        return r

    # Prefer a sheet named Input; else the first sheet with a usable header.
    sheets = list(wb.worksheets)
    chosen = next((s for s in sheets if s.title.strip().lower() == "input"),
                  sheets[0] if sheets else None)
    if chosen is None:
        r = ParseResult()
        r.errors.append("The workbook has no sheets.")
        return r

    rows = [list(r) for r in chosen.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        r = ParseResult()
        r.errors.append(f"Sheet {chosen.title!r} is empty.")
        return r

    # Tolerate a title/legend block above the real header. Pick the row in the
    # first 15 that looks most like a header (most strictly-recognised columns,
    # and must include a name column); earliest wins a tie.
    header_idx, best = 0, -1
    for i, row in enumerate(rows[:15]):
        if not any(canon_header(c, strict=True) == "name" for c in row):
            continue
        sc = score_header_row(row)
        if sc > best:
            best, header_idx = sc, i
    if best < 0:
        # no strict hit anywhere - fall back to a loose scan so decorated
        # headers like "Talent Name (English)" still work
        for i, row in enumerate(rows[:15]):
            if any(canon_header(c) == "name" for c in row):
                header_idx = i
                break
    return _build_rows(rows[header_idx], rows[header_idx + 1:])


def parse_upload(filename: str, data: bytes) -> ParseResult:
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm", ".xltx")):
        return parse_xlsx(data)
    if lower.endswith((".csv", ".tsv", ".txt")):
        return parse_csv(data)
    # sniff: xlsx is a zip
    if data[:2] == b"PK":
        return parse_xlsx(data)
    return parse_csv(data)


# ---------------------------------------------------------------------------
# running a batch
# ---------------------------------------------------------------------------

def flatten_result(pr: ParsedRow, result: dict) -> dict:
    """One resolver result -> one output row."""
    out = {c: "" for c in OUTPUT_COLUMNS}
    out["row_id"] = pr.row_id
    out["name"] = pr.name
    out["notes"] = "; ".join([pr.notes] + pr.warnings).strip("; ")
    decision = result.get("decision", "error")
    out["decision"] = decision

    if decision == "resolved":
        out["entity_id"] = result.get("entity_id", "")
        out["label"] = result.get("label", "")
        out["roles"] = ", ".join(result.get("roles") or [])
        out["coverage"] = result.get("coverage", "")
        handles = result.get("handles") or {}
        for plat in PLATFORM_ORDER:
            hs = handles.get(plat.value) or []
            # multiple wikipedia languages are normal; join them
            out[plat.value] = " | ".join(
                f"{h['handle']}" + (f" ({h['lang']})" if h.get("lang") else "")
                for h in hs)
        out["needs_review"] = " | ".join(
            f"{r['platform']}:{r['handle']}@{r['confidence']}"
            for r in (result.get("needs_review") or []))
        ids = result.get("external_ids") or {}
        out["spoke_ids"] = " | ".join(f"{k}={v}" for k, v in sorted(ids.items()))

    elif decision == "disambiguate":
        out["alternates"] = " | ".join(
            f"{c['entity_id']}:{','.join(c.get('roles') or []) or '?'}"
            f":{(c.get('description') or '')[:40]}"
            for c in (result.get("cards") or []))
        out["notes"] = "; ".join(filter(None, [out["notes"],
                                              result.get("reason", "")]))

    elif decision == "unverified_only":
        out["alternates"] = " | ".join(
            f"{p['platform']}:{p['handle']}"
            for p in (result.get("proposals") or [])[:12])
        out["notes"] = "; ".join(filter(None, [
            out["notes"],
            "search proposals only - UNVERIFIED, confirm before use"]))

    elif decision == "not_found":
        out["notes"] = "; ".join(filter(None, [
            out["notes"], "no matching human in the identity graph"]))
    else:
        out["notes"] = "; ".join(filter(None, [
            out["notes"], str(result.get("error") or "error")]))
    return out


@dataclass
class BatchResult:
    rows: list[dict] = field(default_factory=list)
    processed: int = 0
    skipped: list[dict] = field(default_factory=list)
    unprocessed: list[dict] = field(default_factory=list)
    seconds: float = 0.0
    stopped_because: str = ""

    def as_dict(self) -> dict:
        return {"columns": OUTPUT_COLUMNS, "rows": self.rows,
                "processed": self.processed, "skipped": self.skipped,
                "unprocessed": self.unprocessed,
                "seconds": round(self.seconds, 1),
                "stopped_because": self.stopped_because}


def run_batch(parsed: ParseResult, resolve_one: Callable[[ParsedRow], dict],
              max_rows: int = 500,
              time_budget_seconds: float = 75.0) -> BatchResult:
    """
    Resolve rows under a wall-clock budget.

    Returns whatever finished plus the rows it did not attempt, so a long sheet
    degrades into 'resubmit the rest' rather than a timeout with nothing to
    show. The HTTP cache makes the follow-up pass much cheaper.
    """
    br = BatchResult()
    started = time.monotonic()

    for pr in parsed.invalid_rows:
        br.skipped.append({"row_number": pr.row_number, "row_id": pr.row_id,
                           "why": "; ".join(pr.warnings) or "empty name"})

    todo = parsed.valid_rows[:max_rows]
    for pr in parsed.valid_rows[max_rows:]:
        br.unprocessed.append({"row_number": pr.row_number, "name": pr.name,
                               "why": f"beyond max_rows={max_rows}"})

    for i, pr in enumerate(todo):
        elapsed = time.monotonic() - started
        if elapsed > time_budget_seconds and i > 0:
            br.stopped_because = (
                f"time budget of {time_budget_seconds:.0f}s reached after "
                f"{i} of {len(todo)} rows - resubmit the remainder")
            for rest in todo[i:]:
                br.unprocessed.append({"row_number": rest.row_number,
                                       "name": rest.name,
                                       "why": "time budget"})
            break
        try:
            result = resolve_one(pr)
        except Exception as exc:                          # noqa: BLE001
            result = {"decision": "error", "error": str(exc)[:200]}
        br.rows.append(flatten_result(pr, result))
        br.processed += 1

    br.seconds = time.monotonic() - started
    return br


# ---------------------------------------------------------------------------
# output writers
# ---------------------------------------------------------------------------

def to_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow(r)
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx(rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    head_fill = PatternFill("solid", fgColor="16243F")
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body = Font(name="Arial", size=10)

    ws.append(OUTPUT_COLUMNS)
    for c in range(1, len(OUTPUT_COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    verdict_fill = {
        "resolved": PatternFill("solid", fgColor="E4F2EC"),
        "disambiguate": PatternFill("solid", fgColor="FDF3E0"),
        "unverified_only": PatternFill("solid", fgColor="FDF3E0"),
        "not_found": PatternFill("solid", fgColor="FBECEB"),
        "error": PatternFill("solid", fgColor="FBECEB"),
    }

    for r in rows:
        ws.append([r.get(c, "") for c in OUTPUT_COLUMNS])
        row_idx = ws.max_row
        fill = verdict_fill.get(str(r.get("decision", "")))
        for c in range(1, len(OUTPUT_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = body
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if fill and OUTPUT_COLUMNS[c - 1] == "decision":
                cell.fill = fill

    widths = {"row_id": 8, "name": 24, "decision": 15, "entity_id": 11,
              "label": 24, "roles": 16, "coverage": 9, "label_match": 14,
              "needs_review": 34, "spoke_ids": 40, "alternates": 44,
              "notes": 34}
    for i, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 18)
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
