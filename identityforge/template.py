"""
Generates the upload template.

Per the house rule for a workbook someone else fills in: a legend naming the
cells to edit, and example rows showing realistic values in the expected
format. Three sheets: Instructions, Input (with examples to delete), Reference.
"""

from __future__ import annotations

import io

from .authorities import OCCUPATION_BUCKETS
from .bulk import OUTPUT_COLUMNS, TEMPLATE_COLUMNS

EXAMPLE_ROWS = [
    ["A. R. Rahman", "musician"],
    ["Michael Jackson", "musician"],
    ["Michael Jackson", "athlete"],
    ["Christopher Nolan", "director"],
    ["Priyanka Chopra", "actor"],
    ["\u0936\u093e\u0939\u0930\u0941\u0916\u093c \u0916\u093c\u093e\u0928", "actor"],
]

NOTES = [
    ("name", "REQUIRED", "The talent's name as you have it. Punctuation, "
     "missing diacritics and initials variants are all handled ('A.R. Rahman', "
     "'AR Rahman', 'Beyonce'). Non-Latin scripts are searched in their own "
     "language, so Devanagari or Tamil spellings work as-is."),
    ("profession", "Strongly recommended", "What the person does. This is the "
     "field that separates same-named people - without it, a shared name comes "
     "back as 'disambiguate' instead of an answer. Common words are mapped "
     "automatically: singer, actress, footballer, influencer all work."),
]

DECISIONS = [
    ("resolved", "One person matched on role. Handles in the platform columns "
     "are accepted (confidence >= 0.75)."),
    ("disambiguate", "Several people share this name and the role hint did not "
     "separate them. The alternates column lists the candidates - pick one and "
     "re-run that row with a tighter role or country."),
    ("unverified_only", "No entity in the identity graph, so search proposals "
     "only. UNVERIFIED - confirm before use. Common for micro-influencers."),
    ("not_found", "No matching human found at all. Check the spelling against "
     "the person's Wikipedia article."),
    ("error", "Something failed on that row; see notes."),
]


def build_template_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    NAVY = "16243F"
    ACCENT = "33449C"
    wb = Workbook()

    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor=NAVY)
    body = Font(name="Arial", size=10)
    bold = Font(name="Arial", bold=True, size=10)
    title = Font(name="Arial", bold=True, size=14, color=NAVY)
    # yellow marks cells the user should fill in
    input_fill = PatternFill("solid", fgColor="FFF9DB")
    example_fill = PatternFill("solid", fgColor="EDF1F7")
    example_font = Font(name="Arial", size=10, italic=True, color="5A6B87")
    thin = Side(style="thin", color="D9E0EA")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    # ---------------- Instructions ----------------
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Talent Social Media Finder - bulk upload template"
    ws["A1"].font = title
    ws["A3"] = ("Fill in the Input sheet, then upload it at "
                "https://talent-social-finder.onrender.com")
    ws["A3"].font = body
    ws["A4"] = ("Two columns only: the talent's name, and what they do. "
                "Delete the six grey example rows before uploading.")
    ws["A4"].font = body

    ws["A6"] = "Columns"
    ws["A6"].font = Font(name="Arial", bold=True, size=11, color=NAVY)
    ws.append([])
    r = 7
    for c, label in enumerate(["Column", "Required?", "What it does"], start=1):
        cell = ws.cell(row=r, column=c, value=label)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = wrap
    for name, req, desc in NOTES:
        r += 1
        for c, val in enumerate([name, req, desc], start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = bold if c == 1 else body
            cell.alignment = wrap
            cell.border = box

    r += 3
    ws.cell(row=r, column=1, value="What the decision column means").font = \
        Font(name="Arial", bold=True, size=11, color=NAVY)
    r += 1
    for c, label in enumerate(["decision", "meaning"], start=1):
        cell = ws.cell(row=r, column=c, value=label)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = wrap
    for dec, meaning in DECISIONS:
        r += 1
        for c, val in enumerate([dec, meaning], start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = bold if c == 1 else body
            cell.alignment = wrap
            cell.border = box

    r += 3
    ws.cell(row=r, column=1, value="Limits").font = \
        Font(name="Arial", bold=True, size=11, color=NAVY)
    for line in [
        "A live run costs roughly 5-10 seconds per name, because upstream rate "
        "limits are respected deliberately.",
        "One upload processes rows under a ~75 second budget and returns the "
        "rest as 'unprocessed'. Resubmit those; the cache makes pass two faster.",
        "For a sheet of hundreds of names, upload in batches of about 20-30, or "
        "run it locally where there is no HTTP timeout.",
        "A name match alone never produces a handle. Rows without corroboration "
        "come back as review or unverified, by design.",
    ]:
        r += 1
        cell = ws.cell(row=r, column=1, value="- " + line)
        cell.font = body
        cell.alignment = wrap

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 86
    for row in ws.iter_rows(min_row=3, max_row=r):
        ws.row_dimensions[row[0].row].height = None

    # ---------------- Input ----------------
    inp = wb.create_sheet("Input")
    inp.append(TEMPLATE_COLUMNS)
    for c in range(1, len(TEMPLATE_COLUMNS) + 1):
        cell = inp.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for ex in EXAMPLE_ROWS:
        inp.append(ex)
        for c in range(1, len(TEMPLATE_COLUMNS) + 1):
            cell = inp.cell(row=inp.max_row, column=c)
            cell.font = example_font
            cell.fill = example_fill
            cell.border = box

    # blank rows the operator fills in, marked yellow
    for _ in range(60):
        inp.append([""] * len(TEMPLATE_COLUMNS))
        for c in range(1, len(TEMPLATE_COLUMNS) + 1):
            cell = inp.cell(row=inp.max_row, column=c)
            cell.font = body
            cell.fill = input_fill
            cell.border = box

    roles = sorted(OCCUPATION_BUCKETS.keys())
    dv = DataValidation(type="list", formula1='"' + ",".join(roles) + '"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Pick a role from the list, or leave blank."
    dv.prompt = "Strongly recommended - it is what resolves same-name collisions."
    inp.add_data_validation(dv)
    dv.add(f"B2:B{inp.max_row}")

    widths = {"name": 34, "profession": 22}
    for i, col in enumerate(TEMPLATE_COLUMNS, start=1):
        inp.column_dimensions[get_column_letter(i)].width = widths.get(col, 20)
    inp.freeze_panes = "A2"

    # ---------------- Reference ----------------
    ref = wb.create_sheet("Reference")
    ref["A1"] = "Valid role values"
    ref["A1"].font = Font(name="Arial", bold=True, size=11, color=NAVY)
    ref["A2"] = ("Use any of these in the profession column. Common synonyms "
                 "are mapped automatically, and an unrecognised value is "
                 "ignored rather than guessed at.")
    ref["A2"].font = body
    r = 4
    ref.cell(row=r, column=1, value="role").font = head_font
    ref.cell(row=r, column=1).fill = head_fill
    ref.cell(row=r, column=2, value="matches occupations like").font = head_font
    ref.cell(row=r, column=2).fill = head_fill
    hints = {
        "actor": "actor, actress, film/TV/voice actor",
        "director": "film director, producer",
        "musician": "singer, composer, songwriter, rapper, band",
        "athlete": "footballer, cricketer, basketball player",
        "creator": "YouTuber, streamer, influencer, content creator",
        "model": "model, fashion model",
        "journalist": "journalist, news anchor",
        "politician": "politician",
        "writer": "author, novelist, screenwriter",
        "executive": "CEO, founder, businessperson",
        "comedian": "comedian, stand-up",
        "chef": "chef, restaurateur",
    }
    for role in roles:
        r += 1
        ref.cell(row=r, column=1, value=role).font = bold
        ref.cell(row=r, column=2, value=hints.get(role, "")).font = body
        ref.cell(row=r, column=1).border = box
        ref.cell(row=r, column=2).border = box

    r += 3
    ref.cell(row=r, column=1, value="Output columns you will get back").font = \
        Font(name="Arial", bold=True, size=11, color=NAVY)
    r += 1
    ref.cell(row=r, column=1, value=", ".join(OUTPUT_COLUMNS)).font = body
    ref.cell(row=r, column=1).alignment = wrap

    ref.column_dimensions["A"].width = 30
    ref.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_csv() -> bytes:
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(TEMPLATE_COLUMNS)
    for ex in EXAMPLE_ROWS:
        w.writerow(ex)
    return buf.getvalue().encode("utf-8-sig")
